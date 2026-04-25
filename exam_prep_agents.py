"""
Exam Prep System (Claude Code CLI 기반, API 키 불필요)
Usage:
  python exam_prep_agents.py [date]             # 시험 대비 전체 정리
  python exam_prep_agents.py preview [date]     # 다음 날 예습 (출 빈도 강조)
  python exam_prep_agents.py lecture <file> [date]  # 당일 강의록 통합 분석
  python exam_prep_agents.py compare <new_jungri> <new_chul> [date_range]  # 주말 업데이트 비교
"""

import os
import sys
import re
import subprocess
import concurrent.futures
from datetime import datetime, date, timedelta

import openpyxl
import markdown
from weasyprint import HTML as WeasyHTML

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASE_DIR = "/home/user/endo2"
TIMETABLE_FILE = os.path.join(BASE_DIR, "2023학년도 1학년 2학기 시간표(안)_231005_공지용.xlsx")
JUNGRI_PDF = os.path.join(BASE_DIR, "족보", "[정리족]내분비학 1차 정리족(2).pdf")
CHUL_PDF   = os.path.join(BASE_DIR, "족보", "[출족]내분비학 1차 출족(2) (1).pdf")
SENIORS_DIR = os.path.join(BASE_DIR, "선배족")

WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

TIMETABLE_SECTIONS = [
    (3, 4, 12),
    (14, 15, 23),
    (25, 26, 34),
]

# ---------------------------------------------------------------------------
# Agent 1: 시간표 파싱 (직접 Python으로 처리)
# ---------------------------------------------------------------------------

def agent_timetable(date_str: str) -> dict:
    print(f"[Agent 1] 시간표 파싱 중... ({date_str})")

    date_str = date_str.strip()
    target: date | None = None

    for fmt in ("%Y-%m-%d", "%m/%d", "%Y/%m/%d"):
        try:
            target = datetime.strptime(date_str, fmt).date()
            break
        except ValueError:
            pass

    if target is None:
        m = re.match(r"(\d{1,2})월\s*(\d{1,2})일", date_str)
        if m:
            target = date(2023, int(m.group(1)), int(m.group(2)))

    if target is None:
        return {"error": f"날짜 형식을 인식할 수 없습니다: {date_str}"}

    wb = openpyxl.load_workbook(TIMETABLE_FILE, data_only=True)
    ws = wb.active

    for section in TIMETABLE_SECTIONS:
        date_row, first_row, last_row = section
        for col in range(1, 50):
            cell = ws.cell(row=date_row, column=col).value
            if isinstance(cell, datetime):
                cell_date = cell.date()
            elif isinstance(cell, date):
                cell_date = cell
            else:
                continue
            if cell_date == target:
                classes = []
                for i, row in enumerate(range(first_row, last_row + 1), 1):
                    v = ws.cell(row=row, column=col).value
                    if v and str(v).strip():
                        classes.append({"period": i, "subject": str(v).strip()})
                result = {
                    "date": target.strftime("%Y-%m-%d"),
                    "weekday": WEEKDAY_KR[target.weekday()],
                    "classes": classes,
                }
                print(f"[Agent 1] 완료. 수업 {len(classes)}개 발견.")
                return result

    return {"error": f"{date_str}에 해당하는 날짜를 시간표에서 찾을 수 없습니다."}


# ---------------------------------------------------------------------------
# Claude CLI 실행 헬퍼
# ---------------------------------------------------------------------------

def run_claude(prompt: str, agent_name: str, timeout: int = 600, allowed_tools: str = "Bash,Read") -> str:
    """claude -p 로 서브에이전트를 실행하고 결과를 반환한다."""
    try:
        tools_args = ["--allowedTools", allowed_tools] if allowed_tools and allowed_tools != "none" else []
        work_dir = "/tmp" if not tools_args else BASE_DIR
        result = subprocess.run(
            [
                "claude",
                "--print",
                *tools_args,
                "--output-format", "text",
            ],
            input=prompt,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=work_dir,
        )
        if result.returncode != 0:
            err = result.stderr[:300] if result.stderr else "(오류 메시지 없음)"
            return f"[{agent_name} 오류] {err}"
        return result.stdout.strip()
    except subprocess.TimeoutExpired:
        return f"[{agent_name} 오류] 타임아웃 ({timeout}초)"
    except FileNotFoundError:
        return f"[{agent_name} 오류] claude CLI를 찾을 수 없습니다. Claude Code가 설치되어 있는지 확인하세요."


# ---------------------------------------------------------------------------
# Agent 2: 정리족
# ---------------------------------------------------------------------------

def agent_jungri(classes: list[dict]) -> str:
    subjects = "\n".join(f"- {c['subject']}" for c in classes)

    prompt = f"""당신은 의과대학 시험 대비 정리족 분석 전문가입니다.

아래 수업들의 내용을 정리족 PDF에서 찾아 상세히 정리하세요.

[수업 목록]
{subjects}

[정리족 파일]
{JUNGRI_PDF}

[작업 순서]
1. Bash 도구로 PDF를 텍스트로 변환하세요:
   pdftotext -layout "{JUNGRI_PDF}" /tmp/jungri.txt
2. Bash로 목차를 확인하세요:
   head -n 200 /tmp/jungri.txt
3. 각 수업별로 Bash grep 또는 Read로 섹션을 찾아 내용을 읽으세요.
4. 아래 형식으로 각 수업을 정리하세요:

## [수업명]
### 핵심 개념
### ⭐ 교수 강조 내용 (P 표시)
### 📌 기출 출제 내용 (出 표시)
### 암기 포인트

내용이 길어도 좋으니 최대한 상세하게 작성하세요."""

    return run_claude(prompt, "정리족 Agent")


# ---------------------------------------------------------------------------
# Agent 3: 출족
# ---------------------------------------------------------------------------

def agent_chul(classes: list[dict]) -> str:
    subjects = "\n".join(f"- {c['subject']}" for c in classes)

    prompt = f"""당신은 의과대학 기출문제 분석 전문가입니다.

아래 수업들의 기출문제를 출족 PDF에서 찾아 분석하세요.

[수업 목록]
{subjects}

[출족 파일]
{CHUL_PDF}

[작업 순서]
1. Bash 도구로 PDF를 텍스트로 변환하세요:
   pdftotext -layout "{CHUL_PDF}" /tmp/chul.txt
2. Bash로 목차를 확인하세요:
   head -n 200 /tmp/chul.txt
3. 각 수업별로 기출문제 섹션을 찾아 읽으세요.
4. 아래 형식으로 정리하세요:

## [수업명]
### 기출문제 (최신순)
[연도] 문제 / 정답 / 해설
### 출제 경향 분석
### 반복 출제 포인트

문제와 해설을 모두 포함하고 최대한 많은 문제를 수록하세요."""

    return run_claude(prompt, "출족 Agent")


# ---------------------------------------------------------------------------
# Agent 4: 강의록
# ---------------------------------------------------------------------------

def agent_gangeui(classes: list[dict]) -> str:
    subjects = "\n".join(f"- {c['subject']}" for c in classes)

    # 정리족/출족 제외한 강의 파일 목록
    excluded = {os.path.basename(JUNGRI_PDF), os.path.basename(CHUL_PDF)}
    lecture_files = [
        os.path.join(BASE_DIR, f)
        for f in os.listdir(BASE_DIR)
        if (f.endswith(".pdf") or f.endswith(".pptx")) and f not in excluded
    ]
    files_str = "\n".join(f"- {f}" for f in lecture_files) if lecture_files else "없음"

    prompt = f"""당신은 의과대학 강의록 분석 전문가입니다.

아래 수업들과 관련된 강의 파일을 읽고 핵심 내용을 정리하세요.

[수업 목록]
{subjects}

[이용 가능한 강의 파일]
{files_str}

[작업 순서]
1. 수업명과 관련된 강의 파일을 찾으세요.
2. PDF 파일은 Bash로 읽으세요:
   pdftotext -layout "파일경로" -
3. 강의록에서 중요한 내용, 교수님이 강조한 부분, 임상 예시를 정리하세요.
4. 정리족/출족에서 다루지 않은 추가 내용도 포함하세요.
5. 강의 파일이 없는 수업은 명시하세요.

각 수업별로 구조화하여 정리하세요."""

    return run_claude(prompt, "강의록 Agent")


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def run_exam_prep(date_str: str) -> None:
    print(f"\n{'='*60}")
    print(f"  시험 대비 에이전트 시작: {date_str}")
    print(f"{'='*60}\n")

    # Agent 1: 시간표 파싱
    timetable = agent_timetable(date_str)

    if "error" in timetable:
        print(f"오류: {timetable['error']}")
        return

    classes = timetable.get("classes", [])
    if not classes:
        print(f"{date_str}에 수업이 없습니다.")
        return

    print(f"\n[{timetable['date']} ({timetable['weekday']}요일)] 수업 목록:")
    for c in classes:
        print(f"  {c['period']}교시: {c['subject']}")
    print()

    # Agent 2, 3, 4: 병렬 실행
    print("에이전트 병렬 실행 중 (정리족 / 출족 / 강의록)...\n")

    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        future_jungri = executor.submit(agent_jungri, classes)
        future_chul = executor.submit(agent_chul, classes)
        future_gangeui = executor.submit(agent_gangeui, classes)

        jungri_result = future_jungri.result()
        print("[Agent 2] 정리족 완료.")
        chul_result = future_chul.result()
        print("[Agent 3] 출족 완료.")
        gangeui_result = future_gangeui.result()
        print("[Agent 4] 강의록 완료.")

    # 결과 저장
    safe_date = date_str.replace("/", "-").replace(" ", "_")
    md_path  = os.path.join(BASE_DIR, "md",  f"exam_prep_{safe_date}.md")
    pdf_path = os.path.join(BASE_DIR, "복습", f"exam_prep_{safe_date}.pdf")
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    subjects_md = "\n".join(
        f"- {c['period']}교시: {c['subject']}" for c in classes
    )

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {timetable['date']} ({timetable['weekday']}요일) 시험 대비\n\n")
        f.write(f"## 수업 목록\n\n{subjects_md}\n\n")
        f.write(f"---\n\n## 정리족 요약\n\n{jungri_result}\n\n")
        f.write(f"---\n\n## 출족 분석\n\n{chul_result}\n\n")
        f.write(f"---\n\n## 강의록 보충\n\n{gangeui_result}\n")

    output_path = md_path
    convert_to_pdf(md_path, pdf_path)
    print(f"\n{'='*60}")
    print(f"  결과 저장 완료: {output_path}")
    print(f"  PDF 생성 완료:  {pdf_path}")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# 공통 헬퍼
# ---------------------------------------------------------------------------

def safe_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|\s]', '_', s).strip('_')


def find_seniors_pdfs() -> tuple[str | None, str | None]:
    """선배족 폴더에서 정리족/출족 PDF를 찾아 반환. 없으면 None."""
    jungri, chul = None, None
    if os.path.isdir(SENIORS_DIR):
        for f in os.listdir(SENIORS_DIR):
            if not f.endswith('.pdf'):
                continue
            if '정리족' in f:
                jungri = os.path.join(SENIORS_DIR, f)
            elif '출족' in f:
                chul = os.path.join(SENIORS_DIR, f)
    return jungri, chul


def detect_subject_from_filename(lecture_path: str) -> str | None:
    """파일명에서 과목명 추출. '231023 6교시 갑상샘 종양_김보현 교수님.pdf' → '갑상샘 종양'"""
    stem = os.path.splitext(os.path.basename(lecture_path))[0]
    m = re.search(r'\d+교시\s+(.+?)(?:_|$)', stem)
    return m.group(1).strip() if m else None


_pdf_text_cache: dict[str, str] = {}


def extract_pdf_text(pdf_path: str) -> str:
    """PDF를 텍스트로 변환 (mtime 기반 디스크 캐시 사용)."""
    if pdf_path in _pdf_text_cache:
        return _pdf_text_cache[pdf_path]
    cache_file = f"/tmp/endo2_{os.path.basename(pdf_path)}.txt"
    if os.path.exists(cache_file) and os.path.getmtime(cache_file) >= os.path.getmtime(pdf_path):
        with open(cache_file, encoding="utf-8") as f:
            text = f.read()
    else:
        print(f"  [PDF 추출 중] {os.path.basename(pdf_path)}...")
        r = subprocess.run(
            ["pdftotext", "-layout", pdf_path, "-"],
            capture_output=True, text=True,
        )
        text = r.stdout
        with open(cache_file, "w", encoding="utf-8") as f:
            f.write(text)
    _pdf_text_cache[pdf_path] = text
    return text


def normalize_subject(raw: str) -> tuple[str, str]:
    """'(내분비학)갑상샘 기능 조절약물-김치대' → ('갑상샘 기능 조절약물', '김치대')"""
    s = re.sub(r'^\([^)]+\)', '', raw).strip()
    parts = s.rsplit('-', 1)
    return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (s.strip(), '')


def find_section_in_pdf(full_text: str, subject: str, professor: str) -> str | None:
    """목차에서 과목/교수 위치를 찾아 해당 섹션만 반환. 못 찾으면 None."""
    toc_area = full_text[:4000]
    toc_entries: dict[str, int] = {}
    for m in re.finditer(r'([^\n·.]+?)\s*[·.]{4,}\s*p\s*[.\s]*(\d+)', toc_area):
        name = m.group(1).strip()
        page_str = re.sub(r'\s+', '', m.group(2))
        if page_str.isdigit():
            toc_entries[name] = int(page_str)

    subject_norm = re.sub(r'\s+', '', subject)
    professor_norm = re.sub(r'\s+', '', professor)
    target_page: int | None = None
    fallback_page: int | None = None

    for entry, page in toc_entries.items():
        entry_norm = re.sub(r'\s+', '', entry)
        if subject_norm and len(subject_norm) > 2 and subject_norm in entry_norm:
            target_page = page
            break
        if professor_norm and professor_norm in entry_norm and fallback_page is None:
            fallback_page = page

    if target_page is None:
        target_page = fallback_page
    if target_page is None:
        return None

    all_pages = sorted(set(toc_entries.values()))
    try:
        idx = all_pages.index(target_page)
    except ValueError:
        return None

    next_page = all_pages[idx + 1] if idx + 1 < len(all_pages) else None

    def _find_marker(text, page_num, search_from=0):
        for offset in range(4):
            pat = re.compile(r'\n[ \t]*-[ \t]*' + str(page_num + offset) + r'[ \t]*-[ \t]*\n')
            m = pat.search(text, search_from)
            if m:
                return m
        return None

    start_m = _find_marker(full_text, target_page)
    if not start_m:
        return None
    start = start_m.end()

    if next_page:
        end_m = _find_marker(full_text, next_page, start)
        end = end_m.start() if end_m else len(full_text)
    else:
        end = len(full_text)

    section = full_text[start:end].strip()
    return section if len(section) > 100 else None


# ---------------------------------------------------------------------------
# Feature 1: 예습 (preview)
# ---------------------------------------------------------------------------

def agent_preview_jungri(classes: list[dict], sections_text: str) -> str:
    subjects = "\n".join(f"- {c['subject']}" for c in classes)

    prompt = f"""당신은 의과대학 예습 도우미입니다. 내일 수업을 빠르게 예습할 수 있도록 핵심만 간결하게 정리하세요.

[내일 수업 목록]
{subjects}

[정리족 추출 텍스트]
아래는 각 수업 해당 섹션의 정리족 내용입니다.

{sections_text}

[출력 형식 — 수업마다]
## [수업명]
- **핵심 키워드**: (3~5개)
- **핵심 개념 요약**: (3~5 bullet, 한 줄씩)
- **교수 강조 포인트**: (P 표시 항목)

예습용이므로 각 수업 15줄 이내로 간결하게."""

    return run_claude(prompt, "예습 정리족 Agent", timeout=600, allowed_tools="none")


def agent_preview_chul(classes: list[dict], sections_text: str) -> str:
    subjects = "\n".join(f"- {c['subject']}" for c in classes)

    prompt = f"""당신은 의과대학 기출 출제 경향 분석가입니다. 내일 수업의 출족 여부를 분석하세요.

⚠️ 가장 중요한 것: 각 주제가 기출에 얼마나 자주 나왔는지입니다.

[내일 수업 목록]
{subjects}

[출족 추출 텍스트]
아래는 각 수업 해당 섹션의 출족 내용입니다.

{sections_text}

[출력 형식 — 수업마다]
## [수업명]
🔥 **출 빈도**: ★★★★☆ (5점 만점) — 총 N회 출제 (YYYY, YYYY, ...)
- 자주 출제된 세부 토픽 1
- 자주 출제된 세부 토픽 2
※ 출족에 없으면 "미출제 — 첫 출제 가능성 주시"로 명시

출 빈도 별점을 반드시 포함하세요."""

    return run_claude(prompt, "예습 출족 Agent", timeout=600, allowed_tools="none")


def run_preview(date_str: str) -> None:
    print(f"\n{'='*60}")
    print(f"  예습 에이전트 시작: {date_str}")
    print(f"{'='*60}\n")

    timetable = agent_timetable(date_str)
    if "error" in timetable:
        print(f"오류: {timetable['error']}")
        return
    classes = timetable.get("classes", [])
    if not classes:
        print(f"{date_str}에 수업이 없습니다.")
        return

    print(f"\n[{timetable['date']} ({timetable['weekday']}요일)] 예습 대상 수업:")
    for c in classes:
        print(f"  {c['period']}교시: {c['subject']}")
    print()

    # 선배족 PDF 확인
    seniors_jungri, seniors_chul = find_seniors_pdfs()
    if not seniors_jungri or not seniors_chul:
        print("⚠️  선배족 폴더에 정리족/출족 PDF가 없습니다.")
        print(f"   → {SENIORS_DIR} 에 파일을 업로드한 뒤 다시 실행해주세요.")
        sys.exit(1)

    # Method A: Python에서 PDF 섹션 사전 추출
    print(f"  [사전 추출] 선배족 정리족: {os.path.basename(seniors_jungri)}")
    print(f"  [사전 추출] 선배족 출족:   {os.path.basename(seniors_chul)}\n")
    jungri_text = extract_pdf_text(seniors_jungri)
    chul_text = extract_pdf_text(seniors_chul)

    jungri_parts: list[str] = []
    chul_parts: list[str] = []
    missing_jungri: list[str] = []
    missing_chul: list[str] = []

    for c in classes:
        subject, professor = normalize_subject(c["subject"])
        label = f"{c['period']}교시 {subject}"

        sec_j = find_section_in_pdf(jungri_text, subject, professor)
        if sec_j:
            jungri_parts.append(f"### {label}\n{sec_j}")
        else:
            missing_jungri.append(label)

        sec_c = find_section_in_pdf(chul_text, subject, professor)
        if sec_c:
            chul_parts.append(f"### {label}\n{sec_c}")
        else:
            missing_chul.append(label)

    if missing_jungri or missing_chul:
        print("\n⚠️  다음 수업의 섹션을 PDF에서 찾을 수 없습니다. 확인 후 다시 실행해주세요.\n")
        if missing_jungri:
            print("  [정리족 미발견]")
            for s in missing_jungri:
                print(f"    - {s}")
        if missing_chul:
            print("\n  [출족 미발견]")
            for s in missing_chul:
                print(f"    - {s}")
        print()
        sys.exit(1)

    jungri_sections_text = "\n\n".join(jungri_parts)
    chul_sections_text = "\n\n".join(chul_parts)

    print("예습 에이전트 병렬 실행 중 (정리족 요약 / 출족 빈도)...\n")
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_jungri = executor.submit(agent_preview_jungri, classes, jungri_sections_text)
        f_chul = executor.submit(agent_preview_chul, classes, chul_sections_text)
        preview_jungri = f_jungri.result()
        print("[예습 Agent] 정리족 완료.")
        preview_chul = f_chul.result()
        print("[예습 Agent] 출족 완료.")

    fname = f"preview_{safe_filename(date_str)}"
    md_path  = os.path.join(BASE_DIR, "md",  f"{fname}.md")
    pdf_path = os.path.join(BASE_DIR, "예습", f"{fname}.pdf")
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    subjects_md = "\n".join(f"- {c['period']}교시: {c['subject']}" for c in classes)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {timetable['date']} ({timetable['weekday']}요일) 예습\n\n")
        f.write(f"## 수업 목록\n\n{subjects_md}\n\n")
        f.write(f"---\n\n## 📋 핵심 개념 요약 (정리족)\n\n{preview_jungri}\n\n")
        f.write(f"---\n\n## 🔥 출족 출제 빈도 분석\n\n{preview_chul}\n")

    convert_to_pdf(md_path, pdf_path)
    print(f"\n{'='*60}")
    print(f"  결과 저장 완료: {md_path}")
    print(f"  PDF 생성 완료:  {pdf_path}")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Feature 2: 당일 강의록 통합 (lecture)
# ---------------------------------------------------------------------------

def agent_lecture_integrated(
    lecture_path: str,
    subject: str,
    classes: list[dict],
    seniors_jungri: str,
    seniors_chul: str,
) -> str:
    subjects_context = "\n".join(f"- {c['subject']}" for c in classes)
    ext = os.path.splitext(lecture_path)[1].lower()

    if ext == ".pdf":
        read_instruction = f'pdftotext -layout "{lecture_path}" /tmp/new_lecture.txt && cat /tmp/new_lecture.txt'
    else:
        read_instruction = (
            f'python3 -c "from pptx import Presentation; prs=Presentation(\'{lecture_path}\'); '
            f'[print(shape.text) for slide in prs.slides for shape in slide.shapes if shape.has_text_frame]"'
            f' > /tmp/new_lecture.txt && cat /tmp/new_lecture.txt'
        )

    prompt = f"""당신은 의과대학 당일 강의록 통합 분석 전문가입니다.

오늘 교수님께서 나눠주신 강의 파일을 선배족 정리족/출족과 비교 분석하세요.

[오늘 수업 과목] {subject}
[오늘 전체 수업 목록]
{subjects_context}

[파일 경로]
- 오늘 강의 파일: {lecture_path}
- 선배족 정리족: {seniors_jungri}
- 선배족 출족: {seniors_chul}

[작업 순서]
1. 강의 파일 읽기: {read_instruction}
2. pdftotext -layout "{seniors_jungri}" /tmp/jungri_lec.txt
3. grep -n "{subject}" /tmp/jungri_lec.txt 로 정리족 섹션 위치 확인 후 읽기
4. pdftotext -layout "{seniors_chul}" /tmp/chul_lec.txt
5. grep -n "{subject}" /tmp/chul_lec.txt 로 출족 섹션 위치 확인 후 읽기
6. 세 자료를 비교 분석하여 아래 형식으로 출력

[출력 형식]

## 강의 파일 핵심 내용
(오늘 강의에서 다룬 주요 내용을 구조화하여 정리)

## 정리족과의 비교
### ✅ 정리족과 일치하는 내용
### 🆕 강의에만 있는 새 내용 (중요!)
### ⚠️ 정리족에 있지만 강의에서 다루지 않은 내용

## 출족 관점: 시험 출제 가능성
### 🔥 이번 강의 내용 중 기출 있는 토픽 (연도 및 문제 유형 포함)
### 💡 출족 고빈도 토픽 — 오늘 강의에서 강조된 것
### ⚡ 강의에서 처음 나온 내용 중 출제 가능성 높은 것

## 오늘 수업 요약 암기 포인트
(시험 직전 5분 복습용, 10개 이내)"""

    return run_claude(prompt, f"강의록 통합 Agent ({subject})", timeout=600)


def run_lecture(lecture_path: str, date_str: str) -> None:
    print(f"\n{'='*60}")
    print(f"  당일 강의록 분석 시작: {os.path.basename(lecture_path)}")
    print(f"{'='*60}\n")

    if not os.path.exists(lecture_path):
        print(f"오류: 강의 파일을 찾을 수 없습니다: {lecture_path}")
        return

    # 선배족 PDF 확인
    seniors_jungri, seniors_chul = find_seniors_pdfs()
    if not seniors_jungri or not seniors_chul:
        print("⚠️  선배족 폴더에 정리족/출족 PDF가 없습니다.")
        print(f"   → {SENIORS_DIR} 에 파일을 업로드한 뒤 다시 실행해주세요.")
        return

    timetable = agent_timetable(date_str)
    if "error" in timetable:
        print(f"오류: {timetable['error']}")
        return
    classes = timetable.get("classes", [])

    subject = detect_subject_from_filename(lecture_path)
    if subject is None:
        if classes:
            subject = classes[0]["subject"]
            print(f"[경고] 파일명에서 과목을 인식할 수 없어 '{subject}'로 설정합니다.")
        else:
            subject = os.path.splitext(os.path.basename(lecture_path))[0]
            print(f"[경고] 시간표에서도 과목을 찾을 수 없어 파일명을 사용합니다.")
    else:
        print(f"[과목 자동 감지] {subject}")

    print(f"  선배족 정리족: {os.path.basename(seniors_jungri)}")
    print(f"  선배족 출족:   {os.path.basename(seniors_chul)}\n")

    result = agent_lecture_integrated(lecture_path, subject, classes, seniors_jungri, seniors_chul)
    print("[강의록 통합 Agent] 완료.")

    fname = f"lecture_{safe_filename(date_str)}_{safe_filename(subject)}"
    md_path  = os.path.join(BASE_DIR, "md",  f"{fname}.md")
    pdf_path = os.path.join(BASE_DIR, "복습", f"{fname}.pdf")
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# 당일 강의록 통합 분석\n\n")
        f.write(f"**날짜**: {timetable.get('date', date_str)} ({timetable.get('weekday', '')}요일)\n")
        f.write(f"**과목**: {subject}\n")
        f.write(f"**강의 파일**: `{os.path.basename(lecture_path)}`\n\n")
        f.write(f"---\n\n{result}\n")

    convert_to_pdf(md_path, pdf_path)
    print(f"\n{'='*60}")
    print(f"  결과 저장 완료: {md_path}")
    print(f"  PDF 생성 완료:  {pdf_path}")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Feature 3: 주말 업데이트 비교 (compare)
# ---------------------------------------------------------------------------

def agent_detect_professor_changes(new_jungri_pdf: str) -> str:
    """구버전/신버전 정리족 목차를 비교해 교수가 바뀐 수업을 감지한다."""
    prompt = f"""두 정리족 파일의 목차에서 수업별 담당 교수님을 추출하고 비교하세요.

[구버전 정리족 (작년)]
{JUNGRI_PDF}

[신버전 정리족 (올해)]
{new_jungri_pdf}

[작업]
1. pdftotext -layout "{JUNGRI_PDF}" /tmp/jungri_old_prof.txt && head -n 150 /tmp/jungri_old_prof.txt
2. pdftotext -layout "{new_jungri_pdf}" /tmp/jungri_new_prof.txt && head -n 150 /tmp/jungri_new_prof.txt
3. 목차의 "교수명 – 수업명" 패턴으로 각 버전의 수업-교수 목록을 추출
4. 두 목록을 비교

[출력 형식 — 반드시 이 형식만 출력, 다른 설명 없이]
SAME: [수업명] | [교수명]
CHANGED: [수업명] | [구버전 교수명] → [신버전 교수명]
NEW: [수업명] | [신버전 교수명]
REMOVED: [수업명] | [구버전 교수명]"""

    return run_claude(prompt, "교수 변경 감지 Agent", timeout=300)


def parse_professor_changes(agent_output: str) -> list[tuple[str, str]]:
    """'CHANGED: 수업명 | 구교수 → 신교수' 라인을 파싱해 [(수업명, 설명)] 리스트 반환."""
    changed = []
    for line in agent_output.splitlines():
        line = line.strip()
        if line.startswith("CHANGED:") or line.startswith("NEW:") or line.startswith("REMOVED:"):
            parts = line.split("|", 1)
            subject_part = parts[0].split(":", 1)[1].strip()
            detail = parts[1].strip() if len(parts) > 1 else ""
            changed.append((subject_part, f"{line.split(':')[0]}: {detail}"))
    return changed


def agent_compare_jungri(new_jungri_pdf: str, date_range: str | None, skip_subjects: list[str]) -> str:
    focus = f"\n⚠️ 특히 {date_range} 주간 해당 내용에 집중하세요." if date_range else ""
    skip_note = (
        f"\n\n⛔ 아래 수업은 교수님이 바뀌어 비교 불가 — 완전히 건너뛰세요:\n"
        + "\n".join(f"- {s}" for s in skip_subjects)
    ) if skip_subjects else ""

    prompt = f"""당신은 의과대학 정리족 버전 비교 전문가입니다.
작년 정리족(구버전)과 올해 새 정리족(신버전)을 비교하여 무엇이 달라졌는지 분석하세요.{focus}{skip_note}

[구버전 정리족 (작년)]
{JUNGRI_PDF}

[신버전 정리족 (올해)]
{new_jungri_pdf}

[작업 순서]
1. pdftotext -layout "{JUNGRI_PDF}" /tmp/jungri_old.txt
2. pdftotext -layout "{new_jungri_pdf}" /tmp/jungri_new.txt
3. head -n 200 /tmp/jungri_old.txt 및 head -n 200 /tmp/jungri_new.txt 로 목차 비교
4. 교수 변경 없는 수업만 챕터/섹션별로 내용 비교

[출력 형식]

## 정리족 버전 비교 분석

### 📗 구조 변화
(챕터 추가/삭제/순서 변경)

### 🆕 신버전에만 있는 새 내용
(섹션명 및 핵심 내용 포함)

### ❌ 구버전에서 삭제된 내용

### 📈 강조도 변화
(더 자세해진 섹션 / 줄어든 섹션)

### ⭐ 시험 대비 시사점
(올해 새로 추가된 내용 중 출제 가능성 높은 것)"""

    return run_claude(prompt, "정리족 비교 Agent", timeout=600)


def agent_compare_chul(new_chul_pdf: str, date_range: str | None, skip_subjects: list[str]) -> str:
    focus = f"\n⚠️ 특히 {date_range} 주간 해당 내용에 집중하세요." if date_range else ""
    skip_note = (
        f"\n\n⛔ 아래 수업은 교수님이 바뀌어 비교 불가 — 완전히 건너뛰세요:\n"
        + "\n".join(f"- {s}" for s in skip_subjects)
    ) if skip_subjects else ""

    prompt = f"""당신은 의과대학 출족 버전 비교 전문가입니다.
작년 출족(구버전)과 올해 새 출족(신버전)을 비교하여 기출 트렌드 변화를 분석하세요.{focus}{skip_note}

[구버전 출족 (작년)]
{CHUL_PDF}

[신버전 출족 (올해)]
{new_chul_pdf}

[작업 순서]
1. pdftotext -layout "{CHUL_PDF}" /tmp/chul_old.txt
2. pdftotext -layout "{new_chul_pdf}" /tmp/chul_new.txt
3. head -n 200 /tmp/chul_old.txt 및 head -n 200 /tmp/chul_new.txt 로 목차 비교
4. 교수 변경 없는 수업만 문제 목록 비교

[출력 형식]

## 출족 버전 비교 분석

### 🆕 신버전에 추가된 문제
(과목별로 분류, 문제 내용 + 정답 포함)

### ❌ 삭제된 문제 (구버전에만 있음)

### 📝 해설 변경된 문제
(무엇이 어떻게 바뀌었는지)

### 📊 출제 경향 변화
(어떤 토픽이 더 많이/적게 다뤄지게 됐는지)

### 🎯 올해 시험 대비 전략 업데이트
(신버전 기준으로 우선 학습해야 할 항목)"""

    return run_claude(prompt, "출족 비교 Agent", timeout=600)


def run_compare(new_jungri_pdf: str, new_chul_pdf: str, date_range: str | None) -> None:
    print(f"\n{'='*60}")
    print(f"  주말 업데이트 비교 시작")
    if date_range:
        print(f"  대상 기간: {date_range}")
    print(f"{'='*60}\n")

    for path, label in [(new_jungri_pdf, "신버전 정리족"), (new_chul_pdf, "신버전 출족")]:
        if not os.path.exists(path):
            print(f"오류: {label} 파일을 찾을 수 없습니다: {path}")
            return

    # 교수 변경 감지 (비교 에이전트 실행 전에 먼저)
    print("[사전 검사] 교수 변경 여부 감지 중...\n")
    prof_output = agent_detect_professor_changes(new_jungri_pdf)
    changed_subjects = parse_professor_changes(prof_output)
    skip_subjects = [s for s, _ in changed_subjects]

    if changed_subjects:
        print("⚠️  교수님이 바뀐 수업이 있습니다 — 해당 수업은 비교에서 제외됩니다:\n")
        for subject, detail in changed_subjects:
            print(f"  🔄 {subject}: {detail}")
        print()
    else:
        print("✅ 교수 변경 없음 — 전체 수업 비교를 진행합니다.\n")

    print("비교 에이전트 병렬 실행 중 (정리족 비교 / 출족 비교)...\n")
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_jungri = executor.submit(agent_compare_jungri, new_jungri_pdf, date_range, skip_subjects)
        f_chul = executor.submit(agent_compare_chul, new_chul_pdf, date_range, skip_subjects)
        compare_jungri = f_jungri.result()
        print("[비교 Agent] 정리족 완료.")
        compare_chul = f_chul.result()
        print("[비교 Agent] 출족 완료.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"compare_{timestamp}"
    md_path  = os.path.join(BASE_DIR, "md",  f"{fname}.md")
    pdf_path = os.path.join(BASE_DIR, "복습", f"{fname}.pdf")
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    range_line = f"**비교 기간**: {date_range}\n" if date_range else ""

    changed_section = ""
    if changed_subjects:
        lines = "\n".join(f"- 🔄 **{s}**: {d}" for s, d in changed_subjects)
        changed_section = f"---\n\n## ⚠️ 교수 변경 — 비교 제외 수업\n\n{lines}\n\n"

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# 주말 업데이트 비교 분석\n\n")
        f.write(f"**생성 시각**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"{range_line}")
        f.write(f"**신버전 정리족**: `{os.path.basename(new_jungri_pdf)}`\n")
        f.write(f"**신버전 출족**: `{os.path.basename(new_chul_pdf)}`\n\n")
        f.write(f"{changed_section}")
        f.write(f"---\n\n{compare_jungri}\n\n")
        f.write(f"---\n\n{compare_chul}\n")

    convert_to_pdf(md_path, pdf_path)
    print(f"\n{'='*60}")
    print(f"  결과 저장 완료: {md_path}")
    print(f"  PDF 생성 완료:  {pdf_path}")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# PDF 변환
# ---------------------------------------------------------------------------

_PDF_CSS = """
@import url('file:///usr/share/fonts/truetype/nanum/NanumGothic.ttf');

* { box-sizing: border-box; }

body {
    font-family: 'NanumGothic', 'NanumBarunGothic', sans-serif;
    font-size: 11pt;
    line-height: 1.6;
    color: #1a1a1a;
    margin: 0;
    padding: 0;
}

@page {
    margin: 18mm 15mm 18mm 15mm;
    @bottom-center {
        content: counter(page) " / " counter(pages);
        font-size: 9pt;
        color: #888;
    }
}

h1 { font-size: 20pt; color: #1a3a5c; border-bottom: 2px solid #1a3a5c;
     padding-bottom: 4pt; margin-top: 0; }
h2 { font-size: 14pt; color: #1a3a5c; border-bottom: 1px solid #c0d0e0;
     padding-bottom: 2pt; margin-top: 16pt; }
h3 { font-size: 12pt; color: #2a5a8c; margin-top: 12pt; }
h4 { font-size: 11pt; color: #3a6a9c; margin-top: 8pt; }

p { margin: 4pt 0 6pt 0; }

ul, ol { margin: 4pt 0 6pt 1.5em; padding: 0; }
li { margin: 2pt 0; }

table {
    border-collapse: collapse;
    width: 100%;
    margin: 8pt 0;
    font-size: 10pt;
}
th {
    background: #1a3a5c;
    color: white;
    padding: 5pt 8pt;
    text-align: left;
}
td {
    border: 1px solid #c0d0e0;
    padding: 4pt 8pt;
}
tr:nth-child(even) td { background: #f0f5fa; }

code {
    background: #f4f4f4;
    padding: 1pt 4pt;
    border-radius: 3pt;
    font-size: 9.5pt;
}
pre {
    background: #f4f4f4;
    padding: 8pt;
    border-left: 3pt solid #1a3a5c;
    overflow-x: auto;
    font-size: 9pt;
    line-height: 1.4;
}

blockquote {
    border-left: 3pt solid #7aabcc;
    margin: 6pt 0;
    padding: 4pt 10pt;
    color: #444;
    background: #f0f7fc;
}

hr { border: none; border-top: 1px solid #c0d0e0; margin: 12pt 0; }

strong { color: #c0392b; }
"""


def convert_to_pdf(md_path: str, pdf_path: str | None = None) -> str:
    """마크다운 파일을 PDF로 변환하고 PDF 경로를 반환한다."""
    if pdf_path is None:
        pdf_path = md_path.replace(".md", ".pdf")
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    with open(md_path, encoding="utf-8") as f:
        md_text = f.read()

    body_html = markdown.markdown(
        md_text,
        extensions=["tables", "fenced_code", "nl2br"],
    )

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<style>{_PDF_CSS}</style>
</head>
<body>{body_html}</body>
</html>"""

    WeasyHTML(string=html, base_url=BASE_DIR).write_pdf(pdf_path)
    return pdf_path


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    args = sys.argv[1:]

    if not args or args[0] not in ("preview", "lecture", "compare"):
        run_exam_prep(args[0] if args else "2023-10-23")

    elif args[0] == "preview":
        d = args[1] if len(args) >= 2 else (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")
        run_preview(d)

    elif args[0] == "lecture":
        if len(args) < 2:
            sys.exit("Usage: python exam_prep_agents.py lecture <강의파일경로> [date]")
        run_lecture(args[1], args[2] if len(args) >= 3 else date.today().strftime("%Y-%m-%d"))

    elif args[0] == "compare":
        if len(args) < 3:
            sys.exit("Usage: python exam_prep_agents.py compare <new_jungri.pdf> <new_chul.pdf> [날짜범위]")
        run_compare(args[1], args[2], args[3] if len(args) >= 4 else None)
